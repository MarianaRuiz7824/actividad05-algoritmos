import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment 

Cadena = "TGTAGTGCAGTGGCGTGATCTTGGCTCACTGCAGCCTCCACCTTAGAGCAATCCTCTTGCCTCATCCTCCCGGGTAGTTGGGACTACATGTGCATGCCACATGCCTGGCTAATTTTTGTATTTTTAGTA"
Indices = {43 : "C", 15 : "A", 100 : "G", 54 : "A", 33 : "A", 19 : "G", 97 : "T", 13 : "A"};
Caracteres = list(Cadena);
Lista = [];

def recorrido(Cadena, Longitud, Recorrido,Indices):
    
    Auxiliar = {};

    for i in range(0, len(Cadena), Recorrido):

        for j in range(i, i + Longitud, 1):

            if j in Indices:
                Auxiliar[j] = Indices[j]; #Almacenar los datos de indices validos a intercambiar

        if(len(Auxiliar) > 0): 
           
            permutacion(len(Auxiliar), Auxiliar);

        Auxiliar.clear();

def permutacion(N, Diccionario): #Reducir coste de memoria de m(2^n) a 2^n, solo me interesa los datos que cambian respecto al array original
                                 #Coste ejecución 2^n * m
  Auxiliar = {}
  Indices = list(Diccionario.keys());

  for i in range(2**N):

    for j in range(N -1, -1, -1):

      Posicion = (2**j); #Tasa Cambio

      if(i == 0):
          
          continue;

      Resultado = (i + Posicion) // Posicion; #Paridad  0 mod 2 = 0
      Paridad = Resultado % 2;

      if(Paridad == 0):
        #Número par, causa positiva.
        Auxiliar.update({ Indices[j] : Diccionario[Indices[j]]})

    if(len(Auxiliar) > 0):

        unico(Auxiliar);
    
    Auxiliar.clear();

def unico(Auxiliar):
   
  Unico = True

  for k in range(len(Lista)):
           
    if(Lista[k] == Auxiliar):
              
        Unico = False;
        continue;

  if(Unico):        

    Temporal = Auxiliar.copy();
    Lista.append(Temporal);

def compararLista(Inicio, Final):
   
  Auxiliar = {};
  Inicio -= 1

  for Inicio in range(Final):
     
     Auxiliar.update(Lista[Inicio])
  
  permutacion(len(Auxiliar), Auxiliar);

def imprimirArchivo():
   
  Contador = 0; 
  Encabezado = ["ID", "Cadena", "Cambio"]
  wb = Workbook()
  Hoja = wb.active

  for i in range(len(Encabezado)):
     
     Hoja.cell(row = 1, column = i + 1, value = Encabezado[i]).alignment = Alignment(horizontal='center');

  for i in range(len(Lista)):

    Cambios = convertir(Lista[i].items());
    Temporal = Caracteres.copy() #Variable global para obtener la lista de caracteres.

    if(len(Cambios) > Contador): #Obtener el string de cambios más grande para adecuar el tamaño del excel
       Contador = len(Cambios);

    for Clave, Valor in Lista[i].items():

      Temporal[Clave] = Valor;
    
    Auxiliar = ''.join(Temporal);
    Hoja.cell(row=i+2, column=1, value = i).alignment = Alignment(horizontal='center');
    Hoja.cell(row = i+2, column = 2).value = Auxiliar;
    Hoja.cell(row = i+2, column = 3).value = Cambios;

  Hoja.column_dimensions['B'].width = len(Cadena) * 1.15
  Hoja.column_dimensions['C'].width = Contador * 0.75
  wb.save("Resultados.xlsx");

def convertir(Dato):

  Texto = "";

  for Clave, Valor in Dato:

    Texto += " {} : {} ,".format(Clave, Valor); 
    
  Texto = Texto[:-1]

  return Texto;
  
recorrido(Cadena, 10, 5, Indices);

compararLista(1, len(Lista));

imprimirArchivo();
