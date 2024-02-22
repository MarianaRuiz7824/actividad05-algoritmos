import sys
import pandas as pd
from openpyxl import Workbook
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog, QMessageBox, QVBoxLayout, QWidget, QTextEdit, QGroupBox
from PyQt5.QtGui import QFont
from openpyxl.styles import Alignment

Indices = {}
Cadena = ""
Lista = []
Caracteres = []

class VentanaPrincipal(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Actividad 04: Edición de cadenas de caracteres")
        self.setGeometry(400, 400, 700, 400)
        layout = QVBoxLayout()
        self.boton = QPushButton("Abrir CSV", self)
        self.boton.clicked.connect(self.abrir_csv)
        layout.addWidget(self.boton)
        
        self.nuevo_boton = QPushButton("Mostrar cambios", self)
        self.nuevo_boton.clicked.connect(self.funcion_nuevo_boton)
        layout.addWidget(self.nuevo_boton)
        
        self.group_box_cadena = QGroupBox("Cadena", self)
        self.group_layout_cadena = QVBoxLayout()
        self.text_edit_cadena = QTextEdit(self)
        self.text_edit_cadena.setReadOnly(True)
        font = QFont('Arial', 12)
        self.text_edit_cadena.setFont(font)
        self.group_layout_cadena.addWidget(self.text_edit_cadena)
        self.group_box_cadena.setLayout(self.group_layout_cadena)
        layout.addWidget(self.group_box_cadena)

        self.group_box_columna_0 = QGroupBox("Indices", self)
        self.group_layout_columna_0 = QVBoxLayout()
        self.text_edit_columna_0 = QTextEdit(self)
        self.text_edit_columna_0.setReadOnly(True)
        self.text_edit_columna_0.setFont(font)
        self.group_layout_columna_0.addWidget(self.text_edit_columna_0)
        self.group_box_columna_0.setLayout(self.group_layout_columna_0)
        layout.addWidget(self.group_box_columna_0)
        
        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

        self.setFont(QFont('Arial', 13))
        self.cadena = None
        self.datos_columna_0 = {}
        self.datos_columna_2 = {}

    def abrir_csv(self):
        global Indices
        global Cadena
        global Caracteres
        
        opciones = QFileDialog.Options()
        nombre_archivo, _ = QFileDialog.getOpenFileName(self, "Abrir archivo CSV", "", "Archivos CSV (*.csv)", options=opciones)
        if nombre_archivo:
            try:
                df = pd.read_csv(nombre_archivo, dtype=str)
                
                fila = 0
                columna = 3
                cadena = df.iloc[fila, columna]
                self.text_edit_cadena.setPlainText(str(cadena))
                Cadena = cadena
                
                Caracteres = list(cadena)

                for index, valor in enumerate(df.iloc[0:, 0]):
                    if pd.notna(valor):
                        key = int(valor)
                        value = df.iloc[index, 2]
                        self.datos_columna_0[key] = value

                Indices = self.datos_columna_0
                self.mostrar_datos_en_qtextedit(self.text_edit_columna_0, self.datos_columna_0)

                QMessageBox.information(self, "Éxito", "Los datos han sido extraídos correctamente !!", QMessageBox.Ok)

                print("Datos de los índices de la columna 0:")
                for key, value in self.datos_columna_0.items():
                    print(f"{key}: {value}")
                    
            except Exception as e:
                QMessageBox.warning(self, "Error", f"No se han podido extraer los datos: {str(e)}", QMessageBox.Ok)

    def mostrar_datos_en_qtextedit(self, text_edit, datos):
        texto = ""
        for key, value in datos.items():
            texto += f"{key} : {value}\n"
        text_edit.setPlainText(texto)

    def obtener_datos_columna_0(self):
        return self.datos_columna_0

    def obtener_cadena(self):
        return self.cadena
    
    def funcion_nuevo_boton(self):
        Longitud = 10
        Recorrido = 5
        self.recorrido(Cadena, Longitud, Recorrido, Indices)
        self.compararLista(1, len(Lista))
        self.imprimirArchivo()

    def recorrido(self, Cadena, Longitud, Recorrido, Indices):
        Auxiliar = {}

        for i in range(0, len(Cadena), Recorrido):
            for j in range(i, i + Longitud, 1):
                if j in Indices:
                    Auxiliar[j] = Indices[j]

            if len(Auxiliar) > 0:
                self.permutacion(len(Auxiliar), Auxiliar)

            Auxiliar.clear()

    def permutacion(self, N, Diccionario): 
        Auxiliar = {}
        Indices = list(Diccionario.keys())

        for i in range(2**N):
            for j in range(N - 1, -1, -1):
                Posicion = (2**j)
                if i == 0:
                    continue

                Resultado = (i + Posicion) // Posicion
                Paridad = Resultado % 2

                if Paridad == 0:
                    Auxiliar.update({Indices[j]: Diccionario[Indices[j]]})

            if len(Auxiliar) > 0:
                self.unico(Auxiliar)

            Auxiliar.clear()

    def unico(self, Auxiliar):
        global Lista
        Unico = True

        for k in range(len(Lista)):
            if Lista[k] == Auxiliar:
                Unico = False
                continue

        if Unico:
            Temporal = Auxiliar.copy()
            Lista.append(Temporal)

    def compararLista(self, Inicio, Final):
        global Lista
        Auxiliar = {}
        Inicio -= 1

        for Inicio in range(Final):
            Auxiliar.update(Lista[Inicio])

        self.permutacion(len(Auxiliar), Auxiliar)

    def imprimirArchivo(self):
        global Lista
        global Caracteres
        global Cadena
        Contador = 0
        Encabezado = ["ID", "Cadena", "Cambio"]
        wb = Workbook()
        Hoja = wb.active

        for i in range(len(Encabezado)):
            Hoja.cell(row=1, column=i + 1, value=Encabezado[i]).alignment = Alignment(horizontal='center')

        for i in range(len(Lista)):
            Cambios = self.convertir(Lista[i].items())
            Temporal = Caracteres.copy()

            if len(Cambios) > Contador:
                Contador = len(Cambios)

            for Clave, Valor in Lista[i].items():
                Temporal[Clave] = Valor

            Auxiliar = ''.join(Temporal)
            Hoja.cell(row=i + 2, column=1, value=i).alignment = Alignment(horizontal='center')
            Hoja.cell(row=i + 2, column=2).value = Auxiliar
            Hoja.cell(row=i + 2, column=3).value = Cambios

        Hoja.column_dimensions['B'].width = len(Cadena) * 1.15
        Hoja.column_dimensions['C'].width = Contador * 0.75
        wb.save("Resultados.xlsx")

    def convertir(self, Dato):
        Texto = ""

        for Clave, Valor in Dato:
            Texto += " {} : {} ,".format(Clave, Valor)

        Texto = Texto[:-1]

        return Texto

if __name__ == "__main__":
    app = QApplication(sys.argv)
    ventana = VentanaPrincipal()
    ventana.show()
    sys.exit(app.exec_())
